from __future__ import annotations

import html
from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.pdf import PdfService
from app.modules.reports.contract import Column, ReportResult

_MONEY_FMT = "#,##0.00"


def _cell_value(column: Column, row: dict):
    value = row.get(column.key)
    if value is None:
        return "" if column.type != "money" else None
    if column.type in ("money", "number") and isinstance(value, (Decimal, int, float)):
        return float(value)
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


# --- Excel ---------------------------------------------------------------


def to_xlsx(result: ReportResult) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = result.title[:31]
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="F1F5F9")

    ws.append([result.title])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    if result.subtitle:
        ws.append([result.subtitle])
    ws.append([])

    header_row = ws.max_row + 1
    ws.append([c.label for c in result.columns])
    for idx, col in enumerate(result.columns, start=1):
        cell = ws.cell(row=header_row, column=idx)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="right" if col.align == "right" else "left")

    def write_row(row: dict, *, strong: bool = False):
        r = ws.max_row + 1
        for idx, col in enumerate(result.columns, start=1):
            cell = ws.cell(row=r, column=idx, value=_cell_value(col, row))
            if col.type == "money":
                cell.number_format = _MONEY_FMT
            if col.align == "right":
                cell.alignment = Alignment(horizontal="right")
            if strong:
                cell.font = bold

    for section in result.sections:
        if section.title:
            r = ws.max_row + 1
            ws.cell(row=r, column=1, value=section.title).font = bold
        for row in section.rows:
            write_row(row)
        if section.subtotal:
            write_row(section.subtotal, strong=True)

    if result.grand_total:
        write_row(result.grand_total, strong=True)

    for idx, col in enumerate(result.columns, start=1):
        width = max(len(col.label) + 2, 16 if col.type == "money" else 20)
        ws.column_dimensions[get_column_letter(idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --- PDF -----------------------------------------------------------------


def _fmt(column: Column, row: dict) -> str:
    value = row.get(column.key)
    if value is None or value == "":
        return ""
    if column.type == "money" and isinstance(value, (Decimal, int, float)):
        return f"{float(value):,.2f}"
    if column.type == "number" and isinstance(value, (Decimal, int, float)):
        return f"{float(value):,g}"
    if isinstance(value, date):
        return value.isoformat()
    return html.escape(str(value))


def _row_html(
    columns: list[Column], row: dict, *, strong: bool = False, title: str | None = None
) -> str:
    weight = "font-weight:600;" if strong else ""
    if title is not None:
        return (
            f'<tr><td colspan="{len(columns)}" '
            f'style="padding:8px 6px 4px;font-weight:700;color:#334155;">{html.escape(title)}</td></tr>'
        )
    cells = ""
    for col in columns:
        align = "right" if col.align == "right" else "left"
        cells += (
            f'<td style="padding:4px 6px;text-align:{align};{weight}'
            f'border-top:1px solid #f1f5f9;">{_fmt(col, row)}</td>'
        )
    return f"<tr>{cells}</tr>"


def to_pdf(result: ReportResult, org_name: str | None = None) -> bytes:
    columns = result.columns
    headers = "".join(
        f'<th style="padding:6px;text-align:{"right" if c.align == "right" else "left"};'
        f'font-size:11px;color:#64748b;border-bottom:2px solid #e2e8f0;">{html.escape(c.label)}</th>'
        for c in columns
    )
    body = ""
    for section in result.sections:
        if section.title:
            body += _row_html(columns, {}, title=section.title)
        for row in section.rows:
            body += _row_html(columns, row)
        if section.subtotal:
            body += _row_html(columns, section.subtotal, strong=True)
    if result.grand_total:
        body += _row_html(columns, result.grand_total, strong=True)

    sub = (
        f'<div style="color:#64748b;font-size:12px;">{html.escape(result.subtitle)}</div>'
        if result.subtitle
        else ""
    )
    org = (
        f'<div style="color:#94a3b8;font-size:11px;">{html.escape(org_name)}</div>'
        if org_name
        else ""
    )
    doc = f"""<!doctype html><html><head><meta charset="utf-8"><style>
      body{{font-family:-apple-system,Segoe UI,Roboto,sans-serif;color:#0f172a;margin:0;}}
      table{{width:100%;border-collapse:collapse;font-size:12px;}}
    </style></head><body>
      <div style="margin-bottom:12px;">
        {org}
        <div style="font-size:18px;font-weight:700;">{html.escape(result.title)}</div>
        {sub}
      </div>
      <table><thead><tr>{headers}</tr></thead><tbody>{body}</tbody></table>
    </body></html>"""
    return PdfService().html_to_pdf(doc, paper="a4")
